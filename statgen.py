import openpyxl
from collections import defaultdict
import graph_gen
import teacher

class statgen:
    def __init__(self, file_path: str, first_column : int, end_column : int):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(filename=file_path)
        self.sheet = self.workbook.active
        self.total_answers = self.sheet.max_row - 1
        self.first_column = first_column
        self.end_column = end_column
        self.general = {}
        self.lecturers = []
        self.seminarists = []
        self.labniks = [] 

    def get_statistic_in_column(self, column_number : int):
        value_counts = {}
        title = self.sheet.cell(1, column_number).value
        # Проход по ячейкам столбца, начиная со второй строки
        for row in self.sheet.iter_rows(min_row=2, min_col=column_number, max_col=column_number, values_only=True):
            cell_value = row[0]  # row[0] — значение ячейки в текущем столбце
            if cell_value is not None: # Игнорируем пустые ячейки
                if cell_value.isnumeric():
                    cell_value = int(cell_value)
                if cell_value in value_counts:
                    value_counts[cell_value] += 1
                else:
                    value_counts[cell_value] = 1

        sorted_value_counts = dict(sorted(value_counts.items()))
        return title, sorted_value_counts

    def parse_columns(self, start_column_number : int, end_column_number : int):
        for i in range(start_column_number, end_column_number + 1):
            title, stat = self.get_statistic_in_column(i)
            self.general[title] = stat

    def parse_and_join_columns(self, start_column_number : int, end_column_number : int):
        accum = {}
        for i in range(start_column_number, end_column_number + 1):
            title, stat = self.get_statistic_in_column(i)
            for key, value in stat.items():
                accum[key] = value
        title = title.split(' / ')[0]
        self.general[title] = accum
   
    def find_teachers_by_type(self, teacher_type : str):
        teacher_column = None
        for i in range(self.first_column, self.end_column + 1):  # Проходим по первой строке (заголовки)
            if self.sheet.cell(1, i).value == teacher_type:
                teacher_column = i  # Получаем букву столбца
                break

        if teacher_column is None:
            raise ValueError(f"Столбец {teacher_type} не найден в первой строке файла.")

        # Сбор уникальных значений из столбца
        unique_teachers = set()
        for row in self.sheet.iter_rows(min_row=2, min_col=teacher_column, max_col=teacher_column, values_only=True):
            if row[0]:  # Проверяем, что ячейка не пустая
                unique_teachers.add(row[0])

        if teacher_type == "Лектор":
            self.lecturers =  list(unique_teachers)
        elif teacher_type == "Семинарист":
            self.seminarists = list(unique_teachers)
        elif teacher_type == "Преподаватель по лабораторным работам":
            self.labniks = list(unique_teachers)

    def find_teacher_columns(self, teacher_type : str):
        teacher_type_column = 0
        comment_column = 0
        for i in range(self.first_column, self.end_column + 1):# sheet[1] — первая строка
            print(self.sheet.cell(1, i).value)
            if self.sheet.cell(1, i).value == teacher_type:
                teacher_type_column = i  # Номер столбца
            elif self.sheet.cell(1, i).value.startswith("Ваш комментарий о") and teacher_type_column != 0:
                comment_column = i  # Номер столбца
                break

        if (teacher_type_column == 0 or comment_column == 0):
            raise Exception(f"Столбец {teacher_type} не найден")
        return teacher_type_column, comment_column

    def parse_and_graph_lecturers(self):
        lecturer_start, lecturer_end = self.find_teacher_columns("Лектор")
        self.find_teachers_by_type("Лектор")

        for name in self.lecturers:
            cur_lecturer = teacher.lecturer(self.sheet, lecturer_start, lecturer_end, name)
            print(name)
            cur_lecturer.parse_marks()
            cur_lecturer.parse_questions()
            cur_lecturer.print_marks()
            cur_lecturer.print_questions()
            i = 0
            #import pdb; pdb.set_trace()
            for key, value in cur_lecturer.marks.items(): 
                graph_gen.generate_graph_numbers(value, self.total_answers, key, f"lecturer-marks-{name}-{i}")
                i += 1

            i = 0
            for key, value in cur_lecturer.questions.items():    
                graph_gen.generate_graph_numbers(value, self.total_answers, key, f"lecturer-questions-{name}-{i}")
                i += 1

    def parse_and_graph_seminarists(self):
        seminarist_start, seminarist_end = self.find_teacher_columns("Семинарист")
        self.find_teachers_by_type("Семинарист")

        for name in self.seminarists:
            cur_seminarist = teacher.seminarist(self.sheet, seminarist_start, seminarist_end, name)
            print(name)
            cur_seminarist.parse_marks()
            cur_seminarist.print_marks()
           # import pdb; pdb.set_trace()
            i = 0
            for key, value in cur_seminarist.marks.items():    
                graph_gen.generate_graph_numbers(value, self.total_answers, key, f"seminarists-marks-{name}-{i}")
                i += 1

    def parse_teachers(self, teacher_type : str):
        TEACHER_TYPES = ["Лектор", "Семинарист", "Преподаватель по лабораторным работам"]
        if teacher_type not in TEACHER_TYPES:
            raise Exception("Неправильный тип преподавателя")

        teachers = self.find_all_teachers(teacher_type)
        teacher_start, teacher_end = self.find_teacher_columns(teacher_type)
