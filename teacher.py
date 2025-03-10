import openpyxl
import re

# Абстрактный класс
class teacher:
    def __init__(self, sheet, start_column : int, end_column : int, name : str):
        self.sheet = sheet
        self.start_column = start_column
        self.end_column = end_column
        self.name = name
        self.questions = {}
        self.marks = {}
        self.comments = []
    
    def parse_questions_impl(self, question : str):
        pass

    def parse_marks(self):
        pass

    def parse_comments(self):
        pass

class lecturer(teacher):
    def parse_questions(self):
        
        questions_statistic = {}

        for column_index in range(self.start_column, self.end_column + 1):
            question_answers = 0
            cur_question = self.sheet.cell(1, column_index).value
            title = "Посещали ли Вы лекции?"
            if cur_question.startswith(title):
                row_number = 2
                for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
                    cell_value = row[0]
                    cur_lecturer_name = self.sheet.cell(row_number, self.start_column).value
                    if (cell_value is not None) and (cur_lecturer_name == self.name):
                        question_answers += 1
                    row_number += 1
                cur_question = cur_question.split(' / ')[1]
                questions_statistic[cur_question] = question_answers
        self.questions[title] = questions_statistic
    
    def print_questions(self):
        for elem in self.questions:
            print(f"{elem} {self.questions[elem]}\n")

    def parse_marks(self):
        
        questions_statistic = {}

        for column_index in range(self.start_column, self.end_column + 1):
            question_answers = 0
            cur_question = self.sheet.cell(1, column_index).value
            title = "Оцените качество преподавания лекций" 
            if cur_question.startswith(title):
                row_number = 2
                for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
                    cell_value = row[0]
                    cur_lecturer_name = self.sheet.cell(row_number, self.start_column).value
                    if (cell_value is not None) and (cur_lecturer_name == self.name):
                        cell_value = int(cell_value)
                        if cell_value in questions_statistic:
                            questions_statistic[cell_value] += 1
                        else:
                            questions_statistic[cell_value] = 1

                    row_number += 1
                cur_question = cur_question.split(' / ')[1]
                self.marks[cur_question] = dict(sorted(questions_statistic.items()))
            questions_statistic.clear()

    def parse_comments(self):
        column_index = self.end_column
        
        row_number = 2
        for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
            cell_value = row[0]
            cur_seminarist_name = self.sheet.cell(row_number, self.start_column).value
            if (cell_value is not None) and (cur_seminarist_name == self.name):
                self.comments.append(cell_value)
            row_number += 1

    def print_marks(self):
        for elem in self.marks:
            print(f"{elem}\n{self.marks[elem]}\n")
    
    def print_comments(self):
        file_name = 'txt/lecture_' + self.name + '.txt'
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write(f"{self.name}\n")
            i = 1
            for elem in self.comments:
                file.write(f"{{\n \t\\noindent \\textbf{{Комментарий №{i}.}} \\\\ \n \t{elem} \\\\ \n}}\n\n")
                i = i + 1
            file.write("=====================================================================================\n")

class seminarist(teacher):
    def parse_marks(self):
        
        questions_statistic = {}

        for column_index in range(self.start_column, self.end_column + 1):
            question_answers = 0
            cur_question = self.sheet.cell(1, column_index).value
            title = "Оцените качество преподавания семинаров" 
            if cur_question.startswith(title):
                row_number = 2
                for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
                    cell_value = row[0]
                    cur_seminarist_name = self.sheet.cell(row_number, self.start_column).value
                    if (cell_value is not None) and (cur_seminarist_name == self.name):
                        cell_value = int(cell_value)
                        if cell_value in questions_statistic:
                            questions_statistic[cell_value] += 1
                        else:
                            questions_statistic[cell_value] = 1

                    row_number += 1
                cur_question = cur_question.split(' / ')[1]
                self.marks[cur_question] = dict(sorted(questions_statistic.items()))
            questions_statistic.clear()

    def parse_comments(self):
        column_index = self.end_column
        
        row_number = 2
        for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
            cell_value = row[0]
            cur_seminarist_name = self.sheet.cell(row_number, self.start_column).value
            if (cell_value is not None) and (cur_seminarist_name == self.name):
                self.comments.append(cell_value)
            row_number += 1

    def print_marks(self):
        for elem in self.marks:
            print(f"{elem}\n{self.marks[elem]}\n")

    def print_comments(self):
        file_name = 'txt/seminarist_' + self.name + '.txt'
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write(f"{self.name}\n")
            i = 1
            for elem in self.comments:
                file.write(f"{{\n \t\\noindent \\textbf{{Комментарий №{i}.}} \\\\ \n \t{elem} \\\\ \n}}\n\n")
                i = i + 1
            file.write("=====================================================================================\n")

class labnik(teacher):
    def parse_marks(self):

        questions_statistic = {}

        for column_index in range(self.start_column, self.end_column + 1):
            question_answers = 0
            cur_question = self.sheet.cell(1, column_index).value
            title = "Оцените качество преподавания лабораторных"
            if cur_question.startswith(title):
                row_number = 2
                for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
                    cell_value = row[0]
                    cur_labnik_name = self.sheet.cell(row_number, self.start_column).value
                    if (cell_value is not None) and (cur_labnik_name == self.name):
                        cell_value = int(cell_value)
                        if cell_value in questions_statistic:
                            questions_statistic[cell_value] += 1
                        else:
                            questions_statistic[cell_value] = 1

                    row_number += 1
                cur_question = cur_question.split(' / ')[1]
                self.marks[cur_question] = dict(sorted(questions_statistic.items()))
            questions_statistic.clear()

    def parse_comments(self):
        column_index = self.end_column
        
        row_number = 2
        for row in self.sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
            cell_value = row[0]
            cur_seminarist_name = self.sheet.cell(row_number, self.start_column).value
            if (cell_value is not None) and (cur_seminarist_name == self.name):
                self.comments.append(cell_value)
            row_number += 1

    def print_marks(self):
        for elem in self.marks:
            print(f"{elem}\n{self.marks[elem]}\n")
    
    def print_comments(self):
        file_name = 'txt/labnik_' + self.name + '.txt'
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write(f"{self.name}\n")
            i = 1
            for elem in self.comments:
                file.write(f"{{\n \t\\noindent \\textbf{{Комментарий №{i}.}} \\\\ \n \t{elem} \\\\ \n}}\n\n")
                i = i + 1
            file.write("=====================================================================================\n")
